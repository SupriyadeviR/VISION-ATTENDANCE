from flask import Flask, render_template, request, redirect, jsonify, send_from_directory
import os
import cv2
import csv
import json
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename

from config import UPLOAD_FOLDER, BACKUP_FOLDER, DATABASE_FILE
from utils.db_utils import init_db, get_db, close_db
from utils.face_utils import (
    load_encodings,
    encode_faces,
    recognize_faces_from_frame,
    mark_attendance
)
from utils.backup_utils import backup_database, list_backups, delete_backup

# =====================================================
# APP CONFIG
# =====================================================
app = Flask(__name__)
app.secret_key = "visionpresence-secret"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 5 * 1024 * 1024  # 5MB

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg"}

# =====================================================
# INIT
# =====================================================
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(BACKUP_FOLDER, exist_ok=True)
init_db()
app.teardown_appcontext(close_db)


# Serve uploaded images from the uploads folder
@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)


# Template helper to get basename of a path (works with Windows/Unix paths)
@app.template_filter('basename')
def basename_filter(path):
    if not path:
        return ""
    return os.path.basename(path)


@app.template_filter('format_time')
def format_time_filter(value):
    """Format a time value for templates.
    Accepts datetime.time/datetime/datetime strings and plain strings.
    Returns a human-readable HH:MM:SS or an empty string when falsy.
    """
    if not value:
        return '—'
    # If already a string, assume it's formatted
    if isinstance(value, str):
        return value
    # If it has strftime, use it
    try:
        return value.strftime('%H:%M:%S')
    except Exception:
        return str(value)

# =====================================================
# HELPERS
# =====================================================
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

def get_overview_data():
    """Get overview statistics"""
    conn = get_db()
    
    # Total registered staff
    total_staff = conn.execute("SELECT COUNT(*) as count FROM staff").fetchone()['count']
    
    # Today's date
    today = datetime.now().strftime("%Y-%m-%d")
    
    # Present today
    today_present = conn.execute(
        "SELECT COUNT(DISTINCT emp_id) as count FROM attendance WHERE attendance_date = ?",
        (today,)
    ).fetchone()['count']
    
    # Absent (registered but not marked today)
    today_absent = total_staff - today_present
    
    # Last 7 days attendance average
    week_start = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    week_data = conn.execute("""
        SELECT COUNT(DISTINCT DATE(attendance_date)) as days, COUNT(DISTINCT emp_id) as presents
        FROM attendance WHERE attendance_date >= ?
    """, (week_start,)).fetchone()
    
    week_attendance = int((week_data['presents'] / (total_staff * 7) * 100)) if total_staff > 0 else 0
    
    # Recent activity (with staff names)
    rows = conn.execute("""
        SELECT a.*, s.name FROM attendance a
        JOIN staff s ON a.emp_id = s.emp_id
        ORDER BY a.attendance_date DESC, a.time_in DESC
        LIMIT 10
    """).fetchall()

    recent_activity = []
    for r in rows:
        ti = r['time_in']
        to = r['time_out']
        ti_str = ti if isinstance(ti, str) else (format_time_filter(ti) if ti else None)
        to_str = to if isinstance(to, str) else (format_time_filter(to) if to else None)
        duration = None
        try:
            if ti and to:
                if isinstance(ti, str):
                    ti_dt = datetime.strptime(f"{r['attendance_date']} {ti}", "%Y-%m-%d %H:%M:%S")
                else:
                    ti_dt = ti
                if isinstance(to, str):
                    to_dt = datetime.strptime(f"{r['attendance_date']} {to}", "%Y-%m-%d %H:%M:%S")
                else:
                    to_dt = to
                duration = round((to_dt - ti_dt).total_seconds() / 3600, 2)
        except Exception:
            duration = None

        recent_activity.append({
            'name': r['name'],
            'attendance_date': r['attendance_date'],
            'time_in': ti_str,
            'time_out': to_str,
            'duration': duration
        })
    
    return {
        'total_staff': total_staff,
        'today_present': today_present,
        'today_absent': today_absent,
        'week_attendance': week_attendance,
        'recent_activity': recent_activity
    }

# =====================================================
# DASHBOARD & OVERVIEW
# =====================================================
@app.route("/")
@app.route("/dashboard")
def dashboard():
    return render_template("dashboard.html")

@app.route("/overview")
def overview():
    data = get_overview_data()
    return render_template("overview.html", **data)

# =====================================================
# REGISTER STAFF (FIXED & SAFE)
# =====================================================
@app.route("/register", methods=["GET", "POST"])
def register_staff():
    if request.method == "POST":
        try:
            name = request.form.get("name")
            email = request.form.get("email")
            phone = request.form.get("phone")
            department = request.form.get("department")
            gender = request.form.get("gender")
            joining_date = request.form.get("joining_date")
            dob = request.form.get("dob")
            native = request.form.get("native")

            # ---------- VALIDATION ----------
            if not name:
                return "Name is required", 400

            if "image" not in request.files:
                return "Image file missing", 400

            image = request.files["image"]

            if image.filename == "":
                return "No image selected", 400

            if not allowed_file(image.filename):
                return "Invalid image format", 400

            # ---------- SAVE IMAGE ----------
            filename = secure_filename(f"{name}_{datetime.now().timestamp()}.jpg")
            image_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)

            image.save(image_path)

            # ---------- DB INSERT ----------
            conn = get_db()
            conn.execute("""
                INSERT INTO staff (
                    name, email, phone, department,
                    gender, joining_date, dob, native, image_path
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                name, email, phone, department,
                gender, joining_date, dob, native, image_path
            ))
            conn.commit()

            # ---------- UPDATE FACE MODEL ----------
            encode_faces()

            return redirect("/staff")

        except Exception as e:
            print("REGISTER ERROR:", e)
            return f"Error registering staff: {e}", 500

    return render_template("register_staff.html")

# =====================================================
# STAFF LIST
# =====================================================
@app.route("/staff")
def staff_list():
    conn = get_db()
    staff = conn.execute("SELECT * FROM staff ORDER BY emp_id DESC").fetchall()
    return render_template("staff_list.html", staff=staff)

@app.route("/delete_staff/<int:emp_id>")
def delete_staff(emp_id):
    conn = get_db()
    staff = conn.execute("SELECT image_path FROM staff WHERE emp_id=?", (emp_id,)).fetchone()

    if staff and staff["image_path"] and os.path.exists(staff["image_path"]):
        os.remove(staff["image_path"])

    conn.execute("DELETE FROM staff WHERE emp_id=?", (emp_id,))
    conn.commit()

    encode_faces()
    return redirect("/staff")

# =====================================================
# BIOMETRIC SCAN
# =====================================================
@app.route("/scan")
def scan():
    conn = get_db()
    today = datetime.now().strftime("%Y-%m-%d")
    rows = conn.execute("""
        SELECT a.*, s.name FROM attendance a
        JOIN staff s ON a.emp_id = s.emp_id
        WHERE a.attendance_date = ?
        ORDER BY a.time_in DESC
    """, (today,)).fetchall()

    # Normalize rows to plain dicts with safe time strings and precomputed duration
    today_attendance = []
    for r in rows:
        ti = r['time_in']
        to = r['time_out']
        # time_in/time_out may be stored as strings (HH:MM:SS) or datetime objects
        ti_str = None
        to_str = None
        duration = None

        if ti:
            ti_str = ti if isinstance(ti, str) else format_time_filter(ti)
        if to:
            to_str = to if isinstance(to, str) else format_time_filter(to)

        # compute duration if both available
        try:
            if ti and to:
                # if strings, parse them using attendance_date
                if isinstance(ti, str):
                    ti_dt = datetime.strptime(f"{r['attendance_date']} {ti}", "%Y-%m-%d %H:%M:%S")
                else:
                    ti_dt = ti

                if isinstance(to, str):
                    to_dt = datetime.strptime(f"{r['attendance_date']} {to}", "%Y-%m-%d %H:%M:%S")
                else:
                    to_dt = to

                duration = round((to_dt - ti_dt).total_seconds() / 3600, 2)
        except Exception:
            duration = None

        today_attendance.append({
            'name': r['name'],
            'time_in': ti_str,
            'time_out': to_str,
            'duration': duration,
            'status': r['status']
        })

    return render_template("biometric_scan.html", today_attendance=today_attendance)

@app.route("/mark_attendance", methods=["POST"])
def mark_attendance_route():
    try:
        if "image" not in request.files:
            return jsonify({"status": "failed", "message": "No image provided"}), 400

        image_file = request.files["image"]
        
        if image_file.filename == "":
            return jsonify({"status": "failed", "message": "No file selected"}), 400

        known = load_encodings()
        
        if not known:
            return jsonify({"status": "failed", "message": "No registered staff. Please register staff first."}), 400

        # Save temp image
        temp_path = os.path.join(UPLOAD_FOLDER, "temp_scan.jpg")
        image_file.save(temp_path)

        try:
            import face_recognition
            frame = face_recognition.load_image_file(temp_path)
            results = recognize_faces_from_frame(frame, known)
        except Exception as load_err:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return jsonify({"status": "failed", "message": f"Image processing error: {str(load_err)}"}), 400

        if results:
            emp_id = results[0]["emp_id"]
            conn = get_db()
            
            today = datetime.now().strftime("%Y-%m-%d")
            now = datetime.now().strftime("%H:%M:%S")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Save scanned image
            scans_folder = os.path.join(UPLOAD_FOLDER, "scans")
            os.makedirs(scans_folder, exist_ok=True)
            scan_filename = f"{emp_id}_{timestamp}.jpg"
            scan_path = os.path.join(scans_folder, scan_filename)
            
            try:
                with open(temp_path, 'rb') as src:
                    with open(scan_path, 'wb') as dst:
                        dst.write(src.read())
            except:
                pass  # Continue even if image save fails
            
            if os.path.exists(temp_path):
                os.remove(temp_path)
            
            # Check if already marked today
            existing = conn.execute(
                "SELECT * FROM attendance WHERE emp_id=? AND attendance_date=?",
                (emp_id, today)
            ).fetchone()

            if existing:
                if existing["time_out"]:
                    # Already checked out
                    return jsonify({
                        "status": "failed",
                        "message": f"Already checked out. See you tomorrow!"
                    })
                else:
                    # Instead of auto-setting time_out, ask client to confirm Time Out
                    return jsonify({
                        "status": "need_timeout_confirm",
                        "message": "Employee already checked in today. Confirm to record Time Out.",
                        "attendance_id": existing["id"],
                        "emp_id": emp_id
                    })
            else:
                # New time_in
                conn.execute(
                    "INSERT INTO attendance (emp_id, attendance_date, time_in, status) VALUES (?, ?, ?, ?)",
                    (emp_id, today, now, "Present")
                )
                conn.commit()
                return jsonify({
                    "status": "success",
                    "message": f"✅ Welcome! Time In recorded."
                })
        else:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            return jsonify({
                "status": "failed",
                "message": "❌ Person not registered. Please register first."
            })

    except Exception as e:
        print(f"MARK ATTENDANCE ERROR: {e}")
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return jsonify({"status": "failed", "message": f"Error: {str(e)}"}), 500

# =====================================================
# ATTENDANCE LOGS
# =====================================================
@app.route("/attendance")
def attendance_logs():
    conn = get_db()
    rows = conn.execute("""
        SELECT a.*, s.name FROM attendance a
        JOIN staff s ON a.emp_id = s.emp_id
        ORDER BY a.attendance_date DESC, a.time_in DESC
    """).fetchall()

    # Normalize logs and compute duration safely
    logs = []
    durations = []
    for r in rows:
        ti = r['time_in']
        to = r['time_out']
        ti_str = ti if isinstance(ti, str) else (format_time_filter(ti) if ti else None)
        to_str = to if isinstance(to, str) else (format_time_filter(to) if to else None)
        duration = None
        try:
            if ti and to:
                if isinstance(ti, str):
                    ti_dt = datetime.strptime(f"{r['attendance_date']} {ti}", "%Y-%m-%d %H:%M:%S")
                else:
                    ti_dt = ti
                if isinstance(to, str):
                    to_dt = datetime.strptime(f"{r['attendance_date']} {to}", "%Y-%m-%d %H:%M:%S")
                else:
                    to_dt = to
                duration = round((to_dt - ti_dt).total_seconds() / 3600, 2)
                durations.append(duration)
        except Exception:
            duration = None

        logs.append({
            'id': r['id'],
            'emp_id': r['emp_id'],
            'name': r['name'],
            'attendance_date': r['attendance_date'],
            'time_in': ti_str,
            'time_out': to_str,
            'duration': duration,
            'status': r['status']
        })

    total_records = len(logs)
    unique_employees = len(set(l['emp_id'] for l in logs))

    avg_attendance = 0
    avg_duration = 0
    if logs:
        present_count = sum(1 for log in logs if log['status'] == 'Present')
        avg_attendance = int((present_count / len(logs)) * 100) if logs else 0
        avg_duration = round(sum(durations) / len(durations), 1) if durations else 0

    return render_template("attendance_logs.html",
                         logs=logs,
                         total_records=total_records,
                         unique_employees=unique_employees,
                         avg_attendance=avg_attendance,
                         avg_duration=avg_duration)


# Endpoint to confirm/time-out an attendance record
@app.route("/confirm_timeout", methods=["POST"])
def confirm_timeout():
    try:
        data = request.get_json() or {}
        attendance_id = data.get("attendance_id")

        if not attendance_id:
            return jsonify({"status": "failed", "message": "Missing attendance_id"}), 400

        conn = get_db()
        now = datetime.now().strftime("%H:%M:%S")

        existing = conn.execute("SELECT * FROM attendance WHERE id=?", (attendance_id,)).fetchone()
        if not existing:
            return jsonify({"status": "failed", "message": "Attendance record not found"}), 404

        if existing["time_out"]:
            return jsonify({"status": "failed", "message": "Time Out already recorded"}), 400

        conn.execute(
            "UPDATE attendance SET time_out=?, status=? WHERE id=?",
            (now, "Present", attendance_id)
        )
        conn.commit()
        return jsonify({"status": "success", "message": "✅ Time Out recorded. Have a good day!"})

    except Exception as e:
        return jsonify({"status": "failed", "message": f"Error: {str(e)}"}), 500

# =====================================================
# EXPORT & BACKUP
# =====================================================
@app.route("/export_attendance")
def export_attendance():
    try:
        conn = get_db()
        logs = conn.execute("""
            SELECT a.id, a.emp_id, s.name, a.attendance_date, 
                   a.time_in, a.time_out, a.status
            FROM attendance a
            JOIN staff s ON a.emp_id = s.emp_id
            ORDER BY a.attendance_date DESC, a.time_in DESC
        """).fetchall()

        format_type = request.args.get('format', 'csv').lower()

        if format_type == 'excel':
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"
            
            # Headers
            headers = ['EMP ID', 'Name', 'Date', 'Time In', 'Time Out', 'Duration (Hours)', 'Status']
            ws.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Data rows
            for log in logs:
                duration = ""
                if log['time_in'] and log['time_out']:
                    duration = round((log['time_out'] - log['time_in']).total_seconds() / 3600, 2)
                
                ws.append([
                    log['emp_id'],
                    log['name'],
                    str(log['attendance_date']),
                    format_time_filter(log['time_in']) if log['time_in'] else '',
                    format_time_filter(log['time_out']) if log['time_out'] else '',
                    duration,
                    log['status']
                ])
            
            # Auto adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 16
            ws.column_dimensions['G'].width = 12
            
            # Save to bytes
            import io
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = app.response_class(
                response=output.getvalue(),
                status=200,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response.headers['Content-Disposition'] = 'attachment; filename=attendance.xlsx'
            return response

        elif format_type == 'json':
            data = []
            for log in logs:
                data.append({
                    'emp_id': log['emp_id'],
                    'name': log['name'],
                    'date': str(log['attendance_date']),
                    'time_in': str(log['time_in']) if log['time_in'] else None,
                    'time_out': str(log['time_out']) if log['time_out'] else None,
                    'status': log['status']
                })
            
            response_data = jsonify(data)
            response_data.headers['Content-Disposition'] = 'attachment; filename=attendance.json'
            return response_data

        else:  # CSV
            import io
            si = io.StringIO()
            writer = csv.writer(si)
            writer.writerow(['EMP ID', 'Name', 'Date', 'Time In', 'Time Out', 'Status'])
            
            for log in logs:
                writer.writerow([
                    log['emp_id'],
                    log['name'],
                    log['attendance_date'],
                    format_time_filter(log['time_in']) if log['time_in'] else '',
                    format_time_filter(log['time_out']) if log['time_out'] else '',
                    log['status']
                ])
            
            output = si.getvalue()
            response = app.response_class(
                response=output,
                status=200,
                mimetype='text/csv'
            )
            response.headers['Content-Disposition'] = 'attachment; filename=attendance.csv'
            return response

    except Exception as e:
        return f"Export error: {e}", 500

@app.route("/storage")
def storage():
    # Calculate database size
    db_size = os.path.getsize(DATABASE_FILE) / (1024 * 1024) if os.path.exists(DATABASE_FILE) else 0
    
    # Get backups info
    backups_list = list_backups()
    total_backups = len(backups_list)
    backups_size = sum(b.get('size', 0) for b in backups_list) / 1024 / 1024 if backups_list else 0
    
    return render_template("storage.html",
                         db_size=round(db_size, 2),
                         total_backups=total_backups,
                         backups_size=round(backups_size, 2),
                         backups=backups_list)

@app.route("/backup_db")
def backup_now():
    try:
        backup_database()
        return redirect("/storage")
    except Exception as e:
        return f"Backup error: {e}", 500

@app.route("/download/<filename>")
def download_backup(filename):
    try:
        return send_from_directory(BACKUP_FOLDER, filename, as_attachment=True)
    except Exception as e:
        return f"Download error: {e}", 404

@app.route("/delete_backup/<filename>")
def delete_backup_file(filename):
    try:
        delete_backup(filename)
        return redirect("/storage")
    except Exception as e:
        return f"Delete error: {e}", 500

# =====================================================
# RUN
# =====================================================
if __name__ == "__main__":
    app.run(debug=True)
